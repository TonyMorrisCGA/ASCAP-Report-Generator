import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re

st.set_page_config(page_title="CGA ASCAP Playlist Converter", layout="wide")
st.title("🎸 Classical Guitar Alive! — ASCAP Playlist Converter")
st.markdown("**New feature:** Paste a PRX episode URL and click **Fetch from PRX** — it will auto-scrape the playlist table and fill everything correctly.")

# === PRX URL Scraper ===
url = st.text_input(
    "PRX Episode URL",
    placeholder="https://exchange.prx.org/pieces/626XXX-..."
)

if st.button("🔄 Fetch & Parse from PRX URL", type="primary"):
    if not url:
        st.error("Please enter a PRX URL first.")
    else:
        try:
            # Add ?m=false in case it's needed for full content
            fetch_url = url if "?" in url else url + "?m=false"
            response = requests.get(fetch_url, timeout=15)
            soup = BeautifulSoup(response.text, "html.parser")

            # Try to find episode metadata
            page_title = soup.find("h1").get_text(strip=True) if soup.find("h1") else ""
            episode_title = page_title or "Classical Guitar Alive! Episode"

            # Try to extract date and duration from visible text
            air_date = ""
            total_duration = ""
            for text in soup.stripped_strings:
                if "July" in text or "2026" in text:
                    if not air_date and ("July" in text or re.search(r"\d{1,2}-\d{1,2}-\d{4}", text)):
                        air_date = text.strip()
                if re.search(r"\d{2}:\d{2}", text) and not total_duration:
                    total_duration = text.strip()

            # Find the Musical Works table
            table = None
            for t in soup.find_all("table"):
                headers = [th.get_text(strip=True).lower() for th in t.find_all("th")]
                if "title" in headers and "artist" in headers and "length" in headers:
                    table = t
                    break

            if not table:
                st.error("Could not find the 'Musical Works' table on this page. Try the manual paste option below.")
            else:
                rows = table.find_all("tr")
                data = []
                for i, row in enumerate(rows[1:], start=1):  # skip header
                    cells = [td.get_text(strip=True) for td in row.find_all("td")]
                    if len(cells) >= 6:
                        title = cells[0]
                        artist = cells[1]
                        album = cells[2]
                        label = cells[3]
                        year = cells[4]
                        length = cells[5]

                        # Smart composer extraction (common pattern: "Composer: Piece")
                        composer = title.split(":")[0].strip() if ":" in title else title

                        recording = f"{album} ({label} {year})".strip(" ()")

                        data.append({
                            "Track #": i,
                            "Title": title,
                            "Composer(s)": composer,
                            "Performer(s)": artist,
                            "Duration (MM:SS)": length,
                            "Recording/Album": recording,
                            "Notes": ""
                        })

                if data:
                    st.session_state["playlist_df"] = pd.DataFrame(data)
                    st.session_state["episode_title"] = episode_title
                    if air_date:
                        st.session_state["air_date"] = air_date
                    if total_duration:
                        st.session_state["total_duration"] = total_duration

                    st.success(f"✅ Successfully scraped {len(data)} tracks from PRX!")
                    st.rerun()
                else:
                    st.warning("Table found but no track data extracted.")

        except Exception as e:
            st.error(f"Error fetching or parsing the page: {str(e)}")

# === Manual / Editable Section (always available) ===
st.divider()
st.subheader("📋 Playlist Tracks (edit or add more rows here)")

if "playlist_df" not in st.session_state:
    st.session_state["playlist_df"] = pd.DataFrame(columns=[
        "Track #", "Title", "Composer(s)", "Performer(s)", 
        "Duration (MM:SS)", "Recording/Album", "Notes"
    ])

edited_df = st.data_editor(
    st.session_state["playlist_df"],
    num_rows="dynamic",
    use_container_width=True,
    key="playlist_editor"
)
st.session_state["playlist_df"] = edited_df

# Metadata inputs (auto-filled when possible)
col1, col2 = st.columns(2)
with col1:
    program_title = st.text_input("Program Title", "Classical Guitar Alive!")
    episode_title = st.text_input("Episode Title", st.session_state.get("episode_title", ""))
with col2:
    air_date = st.text_input("Air Date", st.session_state.get("air_date", "July 20, 2026"))
    total_duration = st.text_input("Total Duration", st.session_state.get("total_duration", "58:57"))

# === Generate Excel ===
if st.button("📥 Generate & Download ASCAP Excel", type="primary"):
    if edited_df.empty or len(edited_df) == 0:
        st.error("The playlist table is empty. Please fetch from PRX or add rows manually.")
    else:
        wb = Workbook()
        ws_info = wb.active
        ws_info.title = "Program Info"
        ws_info['A1'] = "ASCAP Music Performance Report"
        ws_info['A1'].font = Font(bold=True, size=14)

        info_rows = [
            ("Program", program_title),
            ("Episode", episode_title),
            ("Air Date", air_date),
            ("Total Duration", total_duration)
        ]
        for idx, (label, value) in enumerate(info_rows, start=3):
            ws_info[f'A{idx}'] = label
            ws_info[f'B{idx}'] = value
            ws_info[f'A{idx}'].font = Font(bold=True)

        # Playlist sheet
        ws_playlist = wb.create_sheet("Playlist")
        headers = ["Track #", "Title", "Composer(s)", "Performer(s)", 
                   "Duration (MM:SS)", "Recording/Album", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws_playlist.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for r_idx, row in enumerate(edited_df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_playlist.cell(row=r_idx, column=c_idx, value=value)

        # Column widths
        for column in ws_playlist.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws_playlist.column_dimensions[column[0].column_letter].width = min(max_length + 3, 80)

        safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', episode_title)[:60]
        filename = f"CGA_{safe_name}_ASCAP.xlsx"
        wb.save(filename)

        with open(filename, "rb") as f:
            st.download_button(
                label="⬇️ Download ASCAP Excel File",
                data=f,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.success("✅ Excel file generated with properly populated data!")
